import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))  #path config.py

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font 


class ReportGenerator:

    def __init__(self, client_name: str, portfolio_value: float, report_date: str = None):
        self.client_name = client_name
        self.portfolio_value = portfolio_value
        self.report_date = report_date or datetime.now().strftime("%Y-%m")
        self.wb = Workbook()

    def add_summary_sheet(self, risk_metrics: dict, optimization_result: dict):
        ws = self.wb.active
        ws.title = "Synthèse"

        # Header
        ws["A1"] = f"Rapport Mensuel — {self.client_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Valeur : {self.portfolio_value:,.0f} € | {self.report_date}"

        # Métriques de risque ✅ VaR et CVaR ajoutés
        ws["A4"] = "MÉTRIQUES DE RISQUE"
        ws["A4"].font = Font(bold=True)
        metrics_map = {
            "Volatilité (%)":      f"{risk_metrics.get('Volatilité (%)', 'N/A'):.2f}",
            "Sharpe Ratio":        str(risk_metrics.get("Sharpe Ratio", "N/A")),
            "Max Drawdown (%)":    f"{risk_metrics.get('Max Drawdown (%)', 'N/A'):.2f}",
            "VaR Param (%)":       str(risk_metrics.get("VaR_param_pct", "N/A")),
            "VaR Hist (%)":        str(risk_metrics.get("VaR_hist_pct", "N/A")),
            "VaR Monte Carlo (%)": str(risk_metrics.get("VaR_mc_pct", "N/A")),
            "CVaR (%)":            str(risk_metrics.get("CVaR_pct", "N/A")),
            "CVaR Monte Carlo (%)":str(risk_metrics.get("CVaR_mc_pct", "N/A")),
        }
        for i, (label, value) in enumerate(metrics_map.items(), start=5):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value

        # Allocation opti
        ws["D4"] = "ALLOCATION OPTIMISÉE"
        ws["D4"].font = Font(bold=True)
        allocations = optimization_result.get("Allocations", {})
        for i, (asset, weight) in enumerate(allocations.items(), start=5):
            ws[f"D{i}"] = asset
            ws[f"E{i}"] = f"{weight:.2f}%"

    def add_stress_test_sheet(self, stress_df: pd.DataFrame): 
        """Onglet dédié aux résultats des stress tests."""
        ws = self.wb.create_sheet(title="Stress Tests")

        ws["A1"] = "STRESS TESTS — SCÉNARIOS DE CRISE"
        ws["A1"].font = Font(bold=True, size=13)

        headers = ["Scénario", "Valeur stressée (€)", "Perte (€)", "Perte (%)"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=header).font = Font(bold=True)

        for row_idx, (scenario, data) in enumerate(stress_df.iterrows(), start=4):
            ws.cell(row=row_idx, column=1, value=scenario)
            ws.cell(row=row_idx, column=2, value=data["Valeur stressée (€)"])
            ws.cell(row=row_idx, column=3, value=data["Perte (€)"])
            ws.cell(row=row_idx, column=4, value=f"{data['Perte (%)']:.2f}%")

    def export_excel(self, filepath: str):
        os.makedirs(os.path.dirname(filepath), exist_ok=True)  
        self.wb.save(filepath)
        print(f"✅ Rapport exporté : {filepath}")